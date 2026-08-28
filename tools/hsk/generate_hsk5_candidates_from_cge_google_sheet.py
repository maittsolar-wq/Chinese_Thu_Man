#!/usr/bin/env python3
"""Build HSK 5 Vietnamese meaning candidates from CGE's public Google Sheet.

The CGE article links a public Google Sheet containing HSK 5 3.0 cumulative
vocabulary with:
Chinese | Pinyin | Part of speech | Vietnamese meaning | example ...

This script downloads that sheet directly as XLSX, so the user does NOT need
to provide any source file.

Input:
    data/hsk/hsk5/hsk5_meanings_candidates_input.json

Output:
    same file

Safety:
- Existing candidateMeanings are preserved.
- Only Vietnamese meanings present in the CGE sheet are added.
- No meanings are invented.
- Base/reviewed/production files are not modified.
"""

from __future__ import annotations

import io
import json
import re
import unicodedata
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "data" / "hsk" / "hsk5"
INPUT = DATA / "hsk5_meanings_candidates_input.json"

SHEET_ID = "1S--Q-63TElZL8Cpg0UXJFzm0lhnrvOjTf8ZbRfjX8LA"
SHEET_URL = (
    f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export"
    "?format=xlsx"
)

EXPECTED = 1600


def download_sheet() -> bytes:
    request = urllib.request.Request(
        SHEET_URL,
        headers={
            "User-Agent": "Mozilla/5.0",
            "Accept-Language": "vi,en;q=0.9",
        },
    )

    with urllib.request.urlopen(
        request,
        timeout=60,
    ) as response:
        return response.read()


def norm_word(value: str) -> str:
    value = str(value or "").strip()
    return re.sub(r"\s+", "", value)


def norm_pinyin(value: str) -> str:
    value = str(value or "").strip().lower()

    value = value.strip("/")
    value = value.replace("’", "")
    value = value.replace("'", "")

    # Remove numbered tones.
    value = re.sub(r"[0-9]", "", value)

    # Remove Unicode tone marks.
    value = unicodedata.normalize(
        "NFD",
        value,
    )

    value = "".join(
        ch
        for ch in value
        if unicodedata.category(ch) != "Mn"
    )

    # Normalize ü/ü to v for matching.
    value = value.replace("ü", "v")

    return re.sub(
        r"[^a-zv]",
        "",
        value,
    )


def clean_meaning(value: str) -> str:
    value = str(value or "").strip()

    value = re.sub(
        r"\s+",
        " ",
        value,
    )

    return value.strip(" ;；|")


def is_chinese_word(value: str) -> bool:
    return bool(
        re.search(
            r"[\u3400-\u9fff]",
            value,
        )
    )


def find_columns(rows):
    """Find the header row and relevant columns."""
    for row_index, row in enumerate(rows):

        values = [
            str(cell or "").strip()
            for cell in row
        ]

        normalized = [
            re.sub(
                r"\s+",
                "",
                value.casefold(),
            )
            for value in values
        ]

        word_col = None
        pinyin_col = None
        meaning_col = None

        for i, value in enumerate(normalized):

            if (
                "từtiếngtrung" in value
                or value == "hántự"
                or value == "từ"
            ):
                word_col = i

            if (
                "phiênâm" in value
                or value == "pinyin"
            ):
                pinyin_col = i

            if (
                "nghĩatiếngviệt" in value
                or value == "nghĩa"
            ):
                meaning_col = i

        if (
            word_col is not None
            and pinyin_col is not None
            and meaning_col is not None
        ):
            return (
                row_index,
                word_col,
                pinyin_col,
                meaning_col,
            )

    raise RuntimeError(
        "Could not locate Chinese/Pinyin/Vietnamese columns "
        "in the downloaded CGE Google Sheet."
    )


def load_reference():
    data = download_sheet()

    workbook = load_workbook(
        io.BytesIO(data),
        read_only=True,
        data_only=True,
    )

    reference = {}
    total_rows = 0

    for sheet in workbook.worksheets:

        rows = sheet.iter_rows(
            values_only=True
        )

        rows = list(rows)

        if not rows:
            continue

        try:
            (
                header_index,
                word_col,
                pinyin_col,
                meaning_col,
            ) = find_columns(rows)
        except RuntimeError:
            continue

        for row in rows[
            header_index + 1:
        ]:

            total_rows += 1

            if max(
                word_col,
                pinyin_col,
                meaning_col,
            ) >= len(row):
                continue

            word = norm_word(
                row[word_col]
            )

            pinyin = norm_pinyin(
                row[pinyin_col]
            )

            meaning = clean_meaning(
                row[meaning_col]
            )

            if not word or not pinyin or not meaning:
                continue

            if not is_chinese_word(word):
                continue

            key = (
                word,
                pinyin,
            )

            reference.setdefault(
                key,
                [],
            )

            if meaning not in reference[key]:
                reference[key].append(
                    meaning
                )

    workbook.close()

    return reference, total_rows


def main():

    print("=" * 72)
    print("HSK 5 MEANING CANDIDATES — CGE GOOGLE SHEET")
    print("=" * 72)
    print()

    if not INPUT.exists():
        raise SystemExit(
            f"Missing input: {INPUT}"
        )

    records = json.loads(
        INPUT.read_text(
            encoding="utf-8"
        )
    )

    if (
        not isinstance(records, list)
        or len(records) != EXPECTED
    ):
        raise SystemExit(
            f"Expected {EXPECTED} candidate records, "
            f"got {len(records)}."
        )

    print(
        "Downloading CGE Google Sheet..."
    )

    try:
        reference, source_rows = (
            load_reference()
        )
    except Exception as exc:
        raise SystemExit(
            "Could not download/read CGE Google Sheet:\n"
            f"{type(exc).__name__}: {exc}"
        )

    print(
        f"Source rows scanned:      "
        f"{source_rows}"
    )

    print(
        f"Reference mappings:       "
        f"{len(reference)}"
    )

    before = 0
    newly = 0
    resolved = 0
    unresolved = []

    for record in records:

        existing = record.get(
            "candidateMeanings",
            [],
        )

        if not isinstance(
            existing,
            list,
        ):
            existing = []

        existing = [
            str(value).strip()
            for value in existing
            if isinstance(
                value,
                str,
            )
            and value.strip()
        ]

        if existing:
            before += 1

        key = (
            norm_word(
                record.get(
                    "word",
                    "",
                )
            ),
            norm_pinyin(
                record.get(
                    "pinyin",
                    "",
                )
            ),
        )

        merged = list(existing)

        seen = {
            value.casefold()
            for value in merged
        }

        for meaning in reference.get(
            key,
            [],
        ):

            normalized = (
                meaning.casefold()
            )

            if normalized not in seen:

                merged.append(
                    meaning
                )

                seen.add(
                    normalized
                )

        if (
            not existing
            and merged
        ):
            newly += 1

        record[
            "candidateMeanings"
        ] = merged

        if merged:

            record[
                "generationStatus"
            ] = (
                "generated_reference_assisted_unverified"
            )

            record[
                "generationSource"
            ] = (
                "cge_google_sheet_hsk30_vietnamese"
            )

            resolved += 1

        else:

            record[
                "generationStatus"
            ] = (
                "needs_manual_verification"
            )

            record[
                "generationSource"
            ] = (
                "no_vietnamese_reference_match"
            )

            unresolved.append(
                record["id"]
            )

    INPUT.write_text(
        json.dumps(
            records,
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    print()
    print("SUCCESS")
    print(
        f"Previously resolved:     "
        f"{before}/{EXPECTED}"
    )

    print(
        f"Newly resolved:          "
        f"{newly}"
    )

    print(
        f"Total resolved:          "
        f"{resolved}/{EXPECTED}"
    )

    print(
        f"Still unresolved:        "
        f"{len(unresolved)}"
    )

    print(
        f"Output:                  "
        f"{INPUT}"
    )

    print()
    print(
        "Existing candidates were preserved."
    )

    print(
        "Only CGE source meanings were added."
    )

    print(
        "No meaning was invented."
    )

    print(
        "Base/reviewed/production data "
        "was not modified."
    )

    if unresolved:

        print()
        print(
            "First unresolved IDs:"
        )

        print(
            ", ".join(
                unresolved[:50]
            )
        )

        if len(unresolved) > 50:
            print(
                f"... and "
                f"{len(unresolved) - 50} more."
            )


if __name__ == "__main__":
    main()
